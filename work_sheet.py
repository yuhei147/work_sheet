import streamlit as st
import openpyxl
import pandas as pd
import datetime
from datetime import datetime
from PIL import Image
import time

num={'向川 有香':'10075','平尾 麗菜':'13061','松原 景一郎':'15060','田中 春希':'14004','近藤 滋義':'14012','生出 翔太':'13002','遠藤 博之':'01036','二浪 佐知子':'03048','谷澤 知恵':'07017','渡辺 康弘':'03040',
    '原田 忠広':'16071','田端 真也':'18043','鈴木 隆信':'16006','藤田 佑磨':'18011','輪違 慶太':'19008','大澤 玄太':'18086','松本 英司':'15007','小笠原 崇文':'17004','若月 大輔':'20014',
    '吉良 智恵':'10073','佐藤 友香':'15047','大野 祐平':'16004','高橋 卓':'17011','松延 怜旺':'19005'}
early_shift=['K','T',"k'","s'","m'",'A','AA']
late_shift=['Q','G','H']
holiday=['公休','有給','特別休暇','FF','RT']
    

st.title('交通費精算/タクシー配車作成くん')
option = st.sidebar.selectbox(
    '作成excel選択',
    ['交通費精算', 'タクシー配車','下準備']
)

if option == '交通費精算':
    st.write('交通費精算')
    uploaded_file=st.file_uploader("csvファイルアップロード",type=['csv'])
    name = st.text_input(
        "名前を入力してください。姓名の間に半角スペース入力してください 👇",)
    if not ' ' in name:
        st.write('姓と名の間に半角スペース入れてください')

    group=st.radio(
        "所属するグループを選んでください 👉",
        ["業務G", "OCG", "運航管理G"],)
    fare = st.text_input(
        "運賃を入力してください👇",)
    if st.button('交通費精算書作成'):
        df=pd.read_csv(uploaded_file)
        df=df.loc[df['名前']==name,:]
        df=df.T
        col=df.columns.values[0]
        df=df.iloc[1:,:]

        data=[]
        for dfs in df[col]:
            data.append(dfs)

        wb = openpyxl.load_workbook('【名前】通勤交通費精算書（2022.6.1.～）.xlsx')
        ws = wb['名前_10月']

        for i, a in zip(range(17,77,2), data):
            ws.unmerge_cells(f'C{i}:C{i+1}')
            ws.cell(row=i,column=3).value=a
            if a in early_shift:
                ws.cell(row=i,column=11).value=1
                ws.cell(row=i+1,column=5).value=1
            elif a in late_shift:
                ws.cell(row=i,column=5).value=1
                ws.cell(row=i+1,column=11).value=1
            elif a in holiday:
                pass
            else:
                ws.cell(row=i,column=5).value=1
                ws.cell(row=i+1,column=5).value=1
        for i in range(17,77,2):
            ws.merge_cells(f'C{i}:C{i+1}')

        ws.cell(row=3,column=5).value=group
        ws.cell(row=8,column=3).value=name
        ws.cell(row=83,column=3).value=fare
        ws.cell(row=12,column=2).value=datetime.now().month
        wb.save(f'【{name}】通勤交通費精算書（2022.6.1.～）.xlsx')
        
        message=st.empty()
        message.write('作成中です')
        st.success('xlsxファイルの出力が完了しました')
        data = open(f'【{name}】通勤交通費精算書（2022.6.1.～）.xlsx', 'rb').read()
        st.download_button(
            label='xlsxダウンロード',
            data=data,
            file_name=f'【{name}】通勤交通費精算書（2022.6.1.～）.xlsx'
        )

elif option=='タクシー配車':
    st.write('タクシー配車')
    uploaded_file=st.file_uploader("csvファイルアップロード",type=['csv'])
    name = st.text_input(
        "名前を入力してください。姓名の間に半角スペース入力してください 👇",)
    if not ' ' in name:
        st.write('姓と名の間に半角スペース入れてください')
    
    adress = st.text_input(
        "住所を入力してください 👇",)

    call_num = st.text_input(
        "電話番号を入力してください 👇",)
    time1= st.text_input(
        "K勤務の配車時間を入力してください 👇",)
    time2= st.text_input(
        "T勤務の配車時間を入力してください 👇",)
    time3= st.text_input(
        "k'勤務の配車時間を入力してください 👇",)
    time4= st.text_input(
        "s'勤務の配車時間を入力してください 👇",)
    time5= st.text_input(
        "m'勤務の配車時間を入力してください 👇",)
    time6= st.text_input(
        "A勤務の配車時間を入力してください 👇",)
    time7= st.text_input(
        "AA勤務の配車時間を入力してください 👇",)
    taxi_time={'K':time1,'T':time2,"k'":time3,"s'":time4,"m'":time5,'A':time6,'AA':time7}

    if st.button('タクシー配車表作成'):
        df=pd.read_csv(uploaded_file)
        df=df.loc[df['名前']==name,:]
        df=df.T
        col=df.columns.values[0]
        df=df.iloc[1:,:]
        wb = openpyxl.load_workbook('タクシー配車表（イースタン）【2022年7月から】 .xlsx')
        ws = wb['TAXI配車表']
        data=[]
        for dfs in df[col]:
            data.append(dfs)
        
        assign_shift=[]
        if 'K'in data:
            assign_shift.append('K')
        if 'T'in data:
            assign_shift.append('T')
        if "k'"in data:
            assign_shift.append("k'")
        if "s'"in data:
            assign_shift.append("s'")
        if "m'"in data:
            assign_shift.append("m'")
        if 'A'in data:
            assign_shift.append('A')
        if 'AA'in data:
            assign_shift.append('AA')

        for j in range(len(assign_shift)):
            ws.cell(row=j+19,column=16).value=assign_shift[j]
            ws.cell(row=j+19,column=17).value=taxi_time[assign_shift[j]]
            ws.cell(row=j+19,column=18).value='下記自宅'
            ws.cell(row=j+19,column=21).value='第2ターミナルビル北玄関'
        
        today = datetime.now()
        today= today.strftime('%Y/%m/%d')
        ws.cell(row=4,column=12).value=today
        now = datetime.now()
        month=str(now.month+1)+'月'
        ws.cell(row=12,column=8).value=mojimoji.han_to_zen(month)

        df_list = []
        for m in range(len(assign_shift)):
            dfs=df[df[col]==assign_shift[m]]
            df_list.append(dfs)
        df1=pd.concat(df_list)
        df1 =df1.sort_index()
        df2=df1.reset_index()
        data2=[]
        for dfss1 in df1[col]:
            data2.append(dfss1)
        data3=[]
        for dfss2 in df2['index']:
            data3.append(dfss2)
        
        for k in range(len(df2)):
            ws.cell(row=k+18,column=2).value=data2[k]
            ws.cell(row=k+18,column=3).value=data3[k]


        ws.cell(row=31,column=3).value=f'自宅： {adress}'

        ws.cell(row=32,column=3).value=f'社員番号：{num[name]} 氏名：{name} 電話：{call_num}'
        wb.save(f'{name} タクシー配車表（イースタン）【2022年7月から】 .xlsx')

        message=st.empty()
        message.write('作成中です')
        st.success('xlsxファイルの出力が完了しました')

        data = open(f'{name} タクシー配車表（イースタン）【2022年7月から】 .xlsx', 'rb').read()
        st.download_button(
            label='xlsxダウンロード',
            data=data,
            file_name=f'{name} タクシー配車表（イースタン）【2022年7月から】 .xlsx'
        )
else:
    st.write('下準備')
    st.write("pdf⇒csv変換はこちら [link](https://www.adobe.com/jp/acrobat/online/pdf-to-excel.html)")

    image = Image.open('ex.png')
    st.image(image, caption='サンプル',use_column_width=True)
    st.write("上記画像のように整形してください（縦に名前、横に日付）")

    st.write("整形できたらアップロードしてください")
    uploaded_file=st.file_uploader("csvファイルアップロード",type=['csv'])
    if st.button('データ変換'):
        df=pd.read_csv(uploaded_file)

        df.replace({r'.*(AA).*': 'AA'}, regex=True, inplace=True)
        df=df.replace('A1','A')
        df=df.replace('A5','A')
        df.replace({r'.*(BA).*': 'BA'}, regex=True, inplace=True)
        df=df.replace('B1','B')
        df=df.replace('B5','B')
        df.replace({r'.*(FA).*': 'FA'}, regex=True, inplace=True)
        df=df.replace('F1','F')
        df=df.replace('F5','F')
        df.replace({r'.*(GA).*': 'GA'}, regex=True, inplace=True)
        df=df.replace('G1','G')
        df=df.replace('G5','G')
        df.replace({r'.*(H).*': 'H'}, regex=True, inplace=True)

        df=df.replace('K1','K')
        df=df.replace('K5','K')
        df=df.replace('T1','T')
        df=df.replace('T5','T')
        df=df.replace('M1','M')
        df=df.replace('M5','M')
        df=df.replace('N1','N')
        df=df.replace('N5','N')
        df=df.replace('U1','U')
        df=df.replace('U5','U')
        df=df.replace('R1','R')
        df=df.replace('R5','R')
        df=df.replace('Q1','Q')
        df=df.replace('Q5','Q')

        df.replace({r".*(k').*": "k'"}, regex=True, inplace=True)
        df.replace({r".*(s').*": "s'"}, regex=True, inplace=True)
        df.replace({r".*(m').*": "m'"}, regex=True, inplace=True)
        df.replace({r".*(f').*": "f'"}, regex=True, inplace=True)
        df.replace({r".*(p').*": "p'"}, regex=True, inplace=True)
        df.replace({r".*(n').*": "n'"}, regex=True, inplace=True)
        df.replace({r".*(r').*": "r'"}, regex=True, inplace=True)
        df.replace({r".*(q').*": "q'"}, regex=True, inplace=True)
        df.replace({r".*(b').*": "b'"}, regex=True, inplace=True)
        df.replace({r".*(c').*": "c'"}, regex=True, inplace=True)
        df.replace({r".*(e').*": "e'"}, regex=True, inplace=True)

        df=df.replace('/', '公休')
        df=df.replace('／', '公休')
        df=df.replace('AL', '有休')
        df=df.replace('D・TR','D')
        df=df.replace('D・BT','D/BT')
        df=df.replace('D・CK','D')
        df=df.replace('特','特別休暇')
        df.replace({r'.*(FF).*': 'FF'}, regex=True, inplace=True)

        df=df.replace('D1','D')
        df=df.replace('D2','D')
        df=df.replace('D3','D')
        df=df.replace('D4','D')
        df=df.replace('D5','D')
        df=df.replace('D6','D')
        time.sleep(15)

        df.to_csv("整形後勤務表.csv", index=False,encoding='utf-8' )

        time.sleep(30)

        data = open(f"整形後勤務表.csv", 'rb').read()
        st.download_button(
            label='xlsxダウンロード',
            data=data,
            file_name="整形後勤務表.csv"
        )

        st.write("文字化けした場合はこちらを参照 [link](https://www.pc-koubou.jp/magazine/38143)")


